from otree.api import *
import csv
import json
import random
import re
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


doc = """
Visual conjoint experiment.

Participants choose between two candidate photos. Candidate metadata is read from:
_static/conjoint/data/dataset.csv

Candidate ideology and policy-focus descriptions are read from:
_static/conjoint/data/ideology_db.xlsx

Candidate photos are stored in:
_static/conjoint/images/

The backend records:
- informed consent
- randomized treatment assignment
- practice versus main-experiment round status
- left/right candidate IDs
- image paths
- all candidate-level metadata from dataset.csv
- candidate-specific ideology and policy-focus descriptions
- full combined candidate metadata rows as JSON backups
- information acquisition behavior
- math-task behavior
- countdown/time-pressure behavior
- timing and mouse-tracking metadata
- post-choice follow-up answers
- Colombia eligibility-screening answers and exclusion reason
- demographic questionnaire answers
- political questionnaire answers
"""


DATASET_COLUMNS = [
    'ID', 'DEPARTAMENTO', 'COD_DPTO', 'MUNICIPIO', 'COD_MCPIO',
    'Cod_candidato', 'NOMBRE', 'Votos', 'PARTIDO', 'COD_PARTIDO',
    'GENERO', 'EDAD', 'age (estimada)', 'gender (estimado)', 'blurness',
    'facequality', 'yaw_angle', 'pitch_angle', 'roll_angle', 'BD',
    'alto', 'ancho', 'fhwr', 'alto_rot', 'ancho_rot', 'fhwr_rot',
    'alto_correg', 'ancho_correg', 'fhwr_correg', 'ranking',
    'partidoFE', 'ideologia_cat', 'fhwr_cat', 'combo_id',
]


FIELD_NAME_MAP = {
    'ID': 'id',
    'DEPARTAMENTO': 'departamento',
    'COD_DPTO': 'cod_dpto',
    'MUNICIPIO': 'municipio',
    'COD_MCPIO': 'cod_mcpio',
    'Cod_candidato': 'cod_candidato',
    'NOMBRE': 'nombre',
    'Votos': 'votos',
    'PARTIDO': 'partido',
    'COD_PARTIDO': 'cod_partido',
    'GENERO': 'genero',
    'EDAD': 'edad',
    'age (estimada)': 'age_estimada',
    'gender (estimado)': 'gender_estimado',
    'blurness': 'blurness',
    'facequality': 'facequality',
    'yaw_angle': 'yaw_angle',
    'pitch_angle': 'pitch_angle',
    'roll_angle': 'roll_angle',
    'BD': 'bd',
    'alto': 'alto',
    'ancho': 'ancho',
    'fhwr': 'fhwr',
    'alto_rot': 'alto_rot',
    'ancho_rot': 'ancho_rot',
    'fhwr_rot': 'fhwr_rot',
    'alto_correg': 'alto_correg',
    'ancho_correg': 'ancho_correg',
    'fhwr_correg': 'fhwr_correg',
    'ranking': 'ranking',
    'partidoFE': 'partido_fe',
    'ideologia_cat': 'ideologia_cat',
    'fhwr_cat': 'fhwr_cat',
    'combo_id': 'combo_id',
}


YES_NO_CHOICES = [
    ['yes', 'Sí'],
    ['no', 'No'],
]

LIVED_IN_COLOMBIA_CHOICES = [
    ['yes', 'Sí'],
    ['no', 'No'],
    ['currently', 'Actualmente vivo en Colombia'],
    ['prefer_not_to_answer', 'Prefiero no responder'],
]

SCREENING_EXCLUSION_LABELS = {
    'residence_colombia': 'Residencia actual en Colombia',
    'nationality_colombian': 'Nacionalidad colombiana',
    'lived_colombia': 'Vivió en Colombia',
    'currently_in_colombia': 'Actualmente vive en Colombia',
    'lived_prefer_not_to_answer': 'No respondió experiencia de residencia',
}

REALISTIC_VOTE_CHOICES = YES_NO_CHOICES

GENDER_CHOICES = [
    ['mujer', 'Mujer'],
    ['hombre', 'Hombre'],
    ['otra', 'Prefiero describirme de otra manera'],
    ['prefiero_no_responder', 'Prefiero no responder'],
]

OCCUPATION_CHOICES = [
    ['tiempo_completo', 'Trabaja remuneradamente a tiempo completo'],
    ['tiempo_parcial', 'Trabaja remuneradamente a tiempo parcial'],
    ['cuenta_propia', 'Trabaja por cuenta propia o de manera independiente'],
    ['estudia', 'Estudia'],
    ['labores_domesticas_cuidado', 'Realiza labores domésticas o de cuidado no remuneradas'],
    ['jubilado_pensionado', 'Está jubilado(a) o pensionado(a)'],
    ['desempleado_busca', 'Está desempleado(a) y buscando trabajo'],
    ['desempleado_no_busca', 'Está desempleado(a) y no busca trabajo actualmente'],
    ['otra', 'Otra situación'],
]

EDUCATION_CHOICES = [
    ['basica_o_menos', 'Educación básica o menos'],
    ['media_completa', 'Educación media completa'],
    [
        'tecnica_profesional_completa',
        'Educación técnica superior o profesional completa',
    ],
    ['postgrado', 'Estudios de postgrado'],
]

POLITICS_FREQUENCY_CHOICES = [
    ['never', 'Nunca'],
    ['less_than_once_per_week', 'Menos de una vez por semana'],
    ['once_or_twice_per_week', 'Una o dos veces por semana'],
    ['several_times_per_week', 'Varias veces por semana'],
    ['every_day', 'Todos los días'],
]

LIKERT_1_TO_7 = [[i, str(i)] for i in range(1, 8)]
LEFT_RIGHT_CHOICES = [[str(i), str(i)] for i in range(0, 11)] + [
    ['no_sabe', 'No sabe'],
    ['no_responde', 'No responde'],
    ['ninguno', 'Ninguno'],
]

EXCEL_ERROR_VALUES = {
    '#NULL!',
    '#DIV/0!',
    '#VALUE!',
    '#REF!',
    '#NAME?',
    '#NUM!',
    '#N/A',
    '#GETTING_DATA',
}


# Stable keys make these bands usable both in the dashboard and in future
# quota rules. Keep the keys unchanged if you later translate the labels.
AGE_BAND_CHOICES = [
    ['18_29', '18–29'],
    ['30_44', '30–44'],
    ['45_59', '45–59'],
    ['60_plus', '60+'],
]


# Fields displayed in the session's admin report. Add future screening fields
# (for example, country_of_residence) here after adding them to Player.
QUOTA_DIMENSIONS = [
    dict(key='age_band', label='Edad', choices=AGE_BAND_CHOICES),
    dict(key='gender_identity', label='Género', choices=GENDER_CHOICES),
    dict(key='education_level', label='Educación', choices=EDUCATION_CHOICES),
    dict(key='occupation_status', label='Ocupación', choices=OCCUPATION_CHOICES),
    dict(
        key='voted_last_municipal',
        label='Votó en la última municipal',
        choices=YES_NO_CHOICES,
    ),
    dict(
        key='political_interest',
        label='Interés político',
        choices=LIKERT_1_TO_7,
    ),
    dict(
        key='politics_frequency',
        label='Frecuencia de conversación política',
        choices=POLITICS_FREQUENCY_CHOICES,
    ),
    dict(
        key='left_right_self_placement',
        label='Ubicación izquierda–derecha',
        choices=LEFT_RIGHT_CHOICES,
    ),
]

QUOTA_REQUIRED_FIELDS = [
    'age_years',
    'gender_identity',
    'education_level',
    'occupation_status',
    'voted_last_municipal',
    'political_interest',
    'politics_frequency',
    'left_right_self_placement',
]


def clean_value(value):
    if value is None:
        return ''
    return str(value).strip()


def clean_candidate_id(value):
    value = clean_value(value)

    if value.endswith('.0'):
        value = value[:-2]

    return Path(value).stem


def find_image_for_candidate(candidate_id, image_dir):
    allowed_extensions = ['.jpg', '.jpeg', '.png', '.webp']

    for ext in allowed_extensions:
        image_path = image_dir / f'{candidate_id}{ext}'
        if image_path.exists():
            return image_path.name

    return None


def load_candidate_ideology_texts(workbook_path):
    if not workbook_path.exists():
        raise FileNotFoundError(
            f'Candidate ideology workbook not found at: {workbook_path}'
        )

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        sheet_name = 'Clasificación Met3'

        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f'{workbook_path.name} must contain a sheet named '
                f'{sheet_name!r}.'
            )

        rows = workbook[sheet_name].iter_rows(values_only=True)

        try:
            headers = [clean_value(value) for value in next(rows)]
        except StopIteration as exc:
            raise ValueError(f'{workbook_path.name} is empty.') from exc

        missing_columns = [
            column for column in ('ID', 'TEXTO')
            if column not in headers
        ]
        if missing_columns:
            raise ValueError(
                f'{workbook_path.name} is missing these expected columns: '
                f'{missing_columns}'
            )

        id_index = headers.index('ID')
        text_index = headers.index('TEXTO')
        ideology_texts = {}
        seen_candidate_ids = set()

        for row_number, row in enumerate(rows, start=2):
            if not any(clean_value(value) for value in row):
                continue

            candidate_id = clean_candidate_id(row[id_index])
            ideology_text = clean_value(row[text_index])

            if candidate_id == '':
                raise ValueError(
                    f'{workbook_path.name} row {row_number} has no candidate ID.'
                )
            if ideology_text == '':
                raise ValueError(
                    f'{workbook_path.name} row {row_number} has no TEXTO for '
                    f'candidate {candidate_id}.'
                )
            if candidate_id in seen_candidate_ids:
                raise ValueError(
                    f'{workbook_path.name} contains duplicate candidate ID '
                    f'{candidate_id}.'
                )

            seen_candidate_ids.add(candidate_id)
            ideology_texts[candidate_id] = (
                None if ideology_text in EXCEL_ERROR_VALUES else ideology_text
            )
    finally:
        workbook.close()

    if not any(ideology_texts.values()):
        raise ValueError(f'{workbook_path.name} contains no candidate text.')

    return ideology_texts


def load_candidate_data():
    project_root = Path(__file__).resolve().parent.parent

    data_path = project_root / '_static' / 'conjoint' / 'data' / 'dataset.csv'
    ideology_path = (
        project_root / '_static' / 'conjoint' / 'data' / 'ideology_db.xlsx'
    )
    image_dir = project_root / '_static' / 'conjoint' / 'images'

    if not data_path.exists():
        raise FileNotFoundError(f'dataset.csv not found at: {data_path}')

    if not image_dir.exists():
        raise FileNotFoundError(f'image folder not found at: {image_dir}')

    ideology_texts = load_candidate_ideology_texts(ideology_path)
    candidates = {}
    missing_ideology_text_ids = []
    invalid_ideology_text_ids = []

    with open(data_path, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)

        if reader.fieldnames is None:
            raise ValueError('dataset.csv has no header row.')

        missing_columns = [
            column for column in DATASET_COLUMNS
            if column not in reader.fieldnames
        ]

        if missing_columns:
            raise ValueError(
                f'dataset.csv is missing these expected columns: {missing_columns}'
            )

        for raw_row in reader:
            row = {
                clean_value(key): clean_value(value)
                for key, value in raw_row.items()
            }

            candidate_id = clean_candidate_id(row.get('ID'))

            if candidate_id == '':
                continue

            image_filename = find_image_for_candidate(candidate_id, image_dir)

            if image_filename is None:
                continue

            if candidate_id not in ideology_texts:
                missing_ideology_text_ids.append(candidate_id)
                continue

            ideology_text = ideology_texts[candidate_id]
            if ideology_text is None:
                invalid_ideology_text_ids.append(candidate_id)
                continue

            row['ID'] = candidate_id
            row['TEXTO'] = ideology_text
            row['_image_filename'] = image_filename
            row['_image_path'] = f'conjoint/images/{image_filename}'

            candidate = {
                'id': candidate_id,
                'image_filename': image_filename,
                'image_path': f'conjoint/images/{image_filename}',
                'ideology_text': ideology_text,
                'dataset_row_json': json.dumps(row, ensure_ascii=False),
            }

            for original_name, field_name in FIELD_NAME_MAP.items():
                candidate[field_name] = clean_value(row.get(original_name))

            candidates[candidate_id] = candidate

    if missing_ideology_text_ids:
        examples = ', '.join(missing_ideology_text_ids[:10])
        raise ValueError(
            'Every randomized candidate must have a TEXTO entry in '
            f'{ideology_path.name}. Missing {len(missing_ideology_text_ids)} '
            f'candidate(s), including: {examples}'
        )

    if invalid_ideology_text_ids:
        print(
            f'Excluded {len(invalid_ideology_text_ids)} candidate(s) from '
            f'randomization because {ideology_path.name} contains an Excel '
            'error instead of TEXTO.'
        )

    if len(candidates) < 2:
        raise ValueError(
            'Need at least two candidate rows with matching image files. '
            f'dataset.csv is at {data_path}. '
            f'Images are expected in {image_dir}.'
        )

    return candidates


CANDIDATE_DATA = load_candidate_data()


class C(BaseConstants):
    NAME_IN_URL = 'conjoint'
    PLAYERS_PER_GROUP = None

    SCREENING_ROUND = 1
    NUM_PRACTICE_ROUNDS = 5
    NUM_MAIN_ROUNDS = 15
    NUM_ROUNDS = NUM_PRACTICE_ROUNDS + NUM_MAIN_ROUNDS

    COUNTDOWN_SECONDS = 10

    TREATMENT_ARMS = [
        'timer_mostrar_mas',
        'captcha_ver_mas',
        'control_ver_mas',
    ]


class Subsession(BaseSubsession):
    def vars_for_admin_report(self):
        return build_admin_report_context(self)


class Group(BaseGroup):
    pass


class Player(BasePlayer):
    consent_accepted = models.BooleanField(initial=False)

    country_of_residence = models.StringField(blank=True)
    nationality = models.StringField(blank=True)
    lived_in_colombia = models.StringField(
        choices=LIVED_IN_COLOMBIA_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )
    screening_excluded = models.BooleanField(initial=False)
    screening_exclusion_reason = models.StringField(blank=True)

    is_practice_round = models.BooleanField(initial=False)
    main_round_number = models.IntegerField(initial=0)

    treatment_arm = models.StringField()
    timed_task = models.BooleanField(initial=False)
    info_condition = models.StringField()
    countdown_seconds = models.IntegerField(initial=C.COUNTDOWN_SECONDS)
    countdown_expired = models.BooleanField(initial=False)

    left_candidate_id = models.StringField()
    right_candidate_id = models.StringField()

    left_image_path = models.StringField()
    right_image_path = models.StringField()

    left_id = models.StringField(blank=True)
    left_departamento = models.StringField(blank=True)
    left_cod_dpto = models.StringField(blank=True)
    left_municipio = models.StringField(blank=True)
    left_cod_mcpio = models.StringField(blank=True)
    left_cod_candidato = models.StringField(blank=True)
    left_nombre = models.StringField(blank=True)
    left_votos = models.StringField(blank=True)
    left_partido = models.StringField(blank=True)
    left_cod_partido = models.StringField(blank=True)
    left_genero = models.StringField(blank=True)
    left_edad = models.StringField(blank=True)
    left_age_estimada = models.StringField(blank=True)
    left_gender_estimado = models.StringField(blank=True)
    left_blurness = models.StringField(blank=True)
    left_facequality = models.StringField(blank=True)
    left_yaw_angle = models.StringField(blank=True)
    left_pitch_angle = models.StringField(blank=True)
    left_roll_angle = models.StringField(blank=True)
    left_bd = models.StringField(blank=True)
    left_alto = models.StringField(blank=True)
    left_ancho = models.StringField(blank=True)
    left_fhwr = models.StringField(blank=True)
    left_alto_rot = models.StringField(blank=True)
    left_ancho_rot = models.StringField(blank=True)
    left_fhwr_rot = models.StringField(blank=True)
    left_alto_correg = models.StringField(blank=True)
    left_ancho_correg = models.StringField(blank=True)
    left_fhwr_correg = models.StringField(blank=True)
    left_ranking = models.StringField(blank=True)
    left_partido_fe = models.StringField(blank=True)
    left_ideologia_cat = models.StringField(blank=True)
    left_ideology_text = models.LongStringField(blank=True)
    left_fhwr_cat = models.StringField(blank=True)
    left_combo_id = models.StringField(blank=True)

    right_id = models.StringField(blank=True)
    right_departamento = models.StringField(blank=True)
    right_cod_dpto = models.StringField(blank=True)
    right_municipio = models.StringField(blank=True)
    right_cod_mcpio = models.StringField(blank=True)
    right_cod_candidato = models.StringField(blank=True)
    right_nombre = models.StringField(blank=True)
    right_votos = models.StringField(blank=True)
    right_partido = models.StringField(blank=True)
    right_cod_partido = models.StringField(blank=True)
    right_genero = models.StringField(blank=True)
    right_edad = models.StringField(blank=True)
    right_age_estimada = models.StringField(blank=True)
    right_gender_estimado = models.StringField(blank=True)
    right_blurness = models.StringField(blank=True)
    right_facequality = models.StringField(blank=True)
    right_yaw_angle = models.StringField(blank=True)
    right_pitch_angle = models.StringField(blank=True)
    right_roll_angle = models.StringField(blank=True)
    right_bd = models.StringField(blank=True)
    right_alto = models.StringField(blank=True)
    right_ancho = models.StringField(blank=True)
    right_fhwr = models.StringField(blank=True)
    right_alto_rot = models.StringField(blank=True)
    right_ancho_rot = models.StringField(blank=True)
    right_fhwr_rot = models.StringField(blank=True)
    right_alto_correg = models.StringField(blank=True)
    right_ancho_correg = models.StringField(blank=True)
    right_fhwr_correg = models.StringField(blank=True)
    right_ranking = models.StringField(blank=True)
    right_partido_fe = models.StringField(blank=True)
    right_ideologia_cat = models.StringField(blank=True)
    right_ideology_text = models.LongStringField(blank=True)
    right_fhwr_cat = models.StringField(blank=True)
    right_combo_id = models.StringField(blank=True)

    left_dataset_row_json = models.LongStringField(blank=True)
    right_dataset_row_json = models.LongStringField(blank=True)

    left_ideology_opened = models.BooleanField(initial=False)
    right_ideology_opened = models.BooleanField(initial=False)

    info_cost_task_completed = models.BooleanField(initial=False)
    info_cost_attempts = models.IntegerField(initial=0)

    captcha_left_a = models.IntegerField(initial=0)
    captcha_left_b = models.IntegerField(initial=0)
    captcha_right_a = models.IntegerField(initial=0)
    captcha_right_b = models.IntegerField(initial=0)

    decision_candidate_id = models.StringField(blank=True)
    decision_side = models.StringField(blank=True)

    time_spent_seconds = models.FloatField(initial=0)
    time_to_first_choice_seconds = models.FloatField(initial=0)

    choice_changes = models.IntegerField(initial=0)
    learn_more_clicks = models.IntegerField(initial=0)

    mouse_distance_px = models.FloatField(initial=0)
    left_hover_seconds = models.FloatField(initial=0)
    right_hover_seconds = models.FloatField(initial=0)

    realistic_vote = models.StringField(
        choices=REALISTIC_VOTE_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )
    decision_factors = models.LongStringField(blank=True)
    rushed_scale = models.IntegerField(
        choices=LIKERT_1_TO_7,
        widget=widgets.RadioSelectHorizontal,
        blank=True,
    )
    info_cost_scale = models.IntegerField(
        choices=LIKERT_1_TO_7,
        widget=widgets.RadioSelectHorizontal,
        blank=True,
    )

    age_years = models.IntegerField(min=18, max=65, blank=True)
    gender_identity = models.StringField(
        choices=GENDER_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )
    gender_identity_other = models.StringField(blank=True)
    occupation_status = models.StringField(
        choices=OCCUPATION_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )
    occupation_status_other = models.StringField(blank=True)
    education_level = models.StringField(
        choices=EDUCATION_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )

    voted_last_municipal = models.StringField(
        choices=YES_NO_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )
    political_interest = models.IntegerField(
        choices=LIKERT_1_TO_7,
        widget=widgets.RadioSelectHorizontal,
        blank=True,
    )
    politics_frequency = models.StringField(
        choices=POLITICS_FREQUENCY_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )
    left_right_self_placement = models.StringField(
        choices=LEFT_RIGHT_CHOICES,
        widget=widgets.RadioSelect,
        blank=True,
    )


def available_candidate_ids():
    return list(CANDIDATE_DATA.keys())


def draw_candidates_for_participant():
    ids = available_candidate_ids()
    needed = C.NUM_ROUNDS * 2

    if len(ids) >= needed:
        return random.sample(ids, needed)

    sampled = []

    while len(sampled) < needed:
        sampled.extend(random.sample(ids, 2))

    return sampled[:needed]


def assign_treatment(player):
    participant = player.participant

    if 'treatment_arm' not in participant.vars:
        participant.vars['treatment_arm'] = random.choice(C.TREATMENT_ARMS)


def creating_session(subsession):
    for player in subsession.get_players():
        participant = player.participant

        assign_treatment(player)

        if 'drawn_candidates' not in participant.vars:
            participant.vars['drawn_candidates'] = draw_candidates_for_participant()


def ensure_participant_vars(player):
    participant = player.participant

    assign_treatment(player)

    if 'drawn_candidates' not in participant.vars:
        participant.vars['drawn_candidates'] = draw_candidates_for_participant()


def save_candidate_to_player(player, side, candidate):
    setattr(player, f'{side}_image_path', candidate['image_path'])
    setattr(player, f'{side}_ideology_text', candidate['ideology_text'])
    setattr(player, f'{side}_dataset_row_json', candidate['dataset_row_json'])

    for field_name in FIELD_NAME_MAP.values():
        setattr(player, f'{side}_{field_name}', candidate.get(field_name, ''))


def assign_candidates(player):
    ensure_participant_vars(player)

    if (
        player.field_maybe_none('left_candidate_id')
        and player.field_maybe_none('right_candidate_id')
    ):
        return

    treatment_arm = player.participant.vars['treatment_arm']

    player.is_practice_round = player.round_number <= C.NUM_PRACTICE_ROUNDS
    player.main_round_number = max(0, player.round_number - C.NUM_PRACTICE_ROUNDS)
    player.treatment_arm = treatment_arm
    player.timed_task = treatment_arm == 'timer_mostrar_mas'
    player.info_condition = treatment_arm
    player.countdown_seconds = C.COUNTDOWN_SECONDS

    if player.captcha_left_a == 0:
        player.captcha_left_a = random.randint(2, 20)
        player.captcha_left_b = random.randint(2, 20)

    if player.captcha_right_a == 0:
        player.captcha_right_a = random.randint(2, 20)
        player.captcha_right_b = random.randint(2, 20)

        while (
            player.captcha_right_a == player.captcha_left_a
            and player.captcha_right_b == player.captcha_left_b
        ):
            player.captcha_right_a = random.randint(2, 20)
            player.captcha_right_b = random.randint(2, 20)

    drawn = player.participant.vars['drawn_candidates']

    start = (player.round_number - 1) * 2
    pair = drawn[start:start + 2]

    if len(pair) < 2:
        pair = random.sample(available_candidate_ids(), 2)

    left_id, right_id = random.sample(pair, 2)

    left = CANDIDATE_DATA[left_id]
    right = CANDIDATE_DATA[right_id]

    player.left_candidate_id = left_id
    player.right_candidate_id = right_id

    save_candidate_to_player(player, 'left', left)
    save_candidate_to_player(player, 'right', right)


def candidate_payload(candidate_id):
    candidate = CANDIDATE_DATA[candidate_id]

    return {
        'id': candidate['id'],
        'image_path': candidate['image_path'],
        'ideology_text': candidate['ideology_text'],
    }


def age_band_key(age):
    if age is None:
        return None
    if age <= 29:
        return '18_29'
    if age <= 44:
        return '30_44'
    if age <= 59:
        return '45_59'
    return '60_plus'


def normalize_screening_text(value):
    value = unicodedata.normalize('NFKD', clean_value(value).casefold())
    return ''.join(
        character
        for character in value
        if not unicodedata.combining(character)
    )


def mentions_colombia(value):
    """Match Colombia/colombiano/colombiana in a free-text response."""
    return bool(re.search(r'\bcolomb', normalize_screening_text(value)))


def screen_out(player, reason):
    player.screening_excluded = True
    player.screening_exclusion_reason = reason
    player.participant.vars['screening_excluded'] = True
    player.participant.vars['screening_exclusion_reason'] = reason


def player_was_screened_out(player):
    stored_on_player = player_field(player, 'screening_excluded')
    return bool(stored_on_player or player.participant.vars.get('screening_excluded'))


def player_is_eligible(player):
    return not player_was_screened_out(player)


def player_field(player, field_name):
    return player.field_maybe_none(field_name)


def profile_is_complete(player):
    return all(
        player_field(player, field) not in (None, '')
        for field in QUOTA_REQUIRED_FIELDS
    )


def has_any_profile_answer(player):
    return any(
        player_field(player, field) not in (None, '')
        for field in QUOTA_REQUIRED_FIELDS
    )


def quota_value(player, dimension_key):
    if dimension_key == 'age_band':
        return age_band_key(player_field(player, 'age_years'))
    return player_field(player, dimension_key)


def quota_target(targets, dimension_key, category_key):
    dimension_targets = targets.get(dimension_key, {})
    target = dimension_targets.get(category_key)
    if isinstance(target, int) and not isinstance(target, bool) and target > 0:
        return target
    return None


def build_quota_dimension(players, dimension, targets):
    values = [quota_value(player, dimension['key']) for player in players]
    counts = Counter(value for value in values if value not in (None, ''))
    answered = sum(counts.values())
    rows = []

    for key, label in dimension['choices']:
        count = counts[key]
        target = quota_target(targets, dimension['key'], key)
        share = round((count / answered) * 100) if answered else 0
        target_progress = round((count / target) * 100) if target else None
        rows.append(
            dict(
                key=key,
                label=label,
                count=count,
                share=share,
                target=target,
                is_full=target is not None and count >= target,
                bar_width=min(target_progress if target_progress is not None else share, 100),
            )
        )

    return dict(
        key=dimension['key'],
        label=dimension['label'],
        answered=answered,
        missing=len(players) - answered,
        rows=rows,
    )


def build_age_gender_intersection(players, targets):
    gender_columns = [dict(key=key, label=label) for key, label in GENDER_CHOICES]
    counts = Counter()

    for player in players:
        age_key = age_band_key(player_field(player, 'age_years'))
        gender_key = player_field(player, 'gender_identity')
        if age_key and gender_key:
            counts[(age_key, gender_key)] += 1

    rows = []
    for age_key, age_label in AGE_BAND_CHOICES:
        cells = []
        for gender in gender_columns:
            count = counts[(age_key, gender['key'])]
            target_key = f"{age_key}__{gender['key']}"
            target = quota_target(targets, 'age_gender', target_key)
            cells.append(
                dict(
                    count=count,
                    target=target,
                    is_full=target is not None and count >= target,
                )
            )
        rows.append(
            dict(
                key=age_key,
                label=age_label,
                cells=cells,
                total=sum(cell['count'] for cell in cells),
            )
        )

    return dict(gender_columns=gender_columns, rows=rows)


def participant_status(participant, complete, screened_out):
    if screened_out:
        return 'Excluido por filtro'
    if complete:
        return 'Perfil completo'
    if not participant.visited:
        return 'No inició'
    if participant._index_in_pages >= participant._max_page_index:
        return 'Página final'
    return 'En curso'


def build_admin_report_context(subsession):
    """Build a one-participant-per-row quota snapshot for the admin report."""
    is_official = bool(
        subsession.session.config.get('official_data_collection', False)
    )
    screening_players = list(
        subsession.in_round(C.SCREENING_ROUND).get_players()
    )

    # Sessions created before the questionnaire was moved stored these fields
    # in the last round. Keep those historical sessions visible in the report.
    if C.SCREENING_ROUND != C.NUM_ROUNDS:
        legacy_players = {
            player.participant.id: player
            for player in subsession.in_round(C.NUM_ROUNDS).get_players()
        }
        players = [
            player
            if has_any_profile_answer(player)
            else legacy_players.get(player.participant.id, player)
            for player in screening_players
        ]
    else:
        players = screening_players

    targets = subsession.session.config.get('quota_targets', {}) or {}

    age_labels = dict(AGE_BAND_CHOICES)
    gender_labels = dict(GENDER_CHOICES)
    participant_rows = []

    for player in players:
        participant = player.participant
        screened_out = player_was_screened_out(player)
        complete = profile_is_complete(player)
        age_key = age_band_key(player_field(player, 'age_years'))
        gender_key = player_field(player, 'gender_identity')
        participant_rows.append(
            dict(
                code=participant.code,
                status=participant_status(participant, complete, screened_out),
                screened_out=screened_out,
                exclusion_reason=SCREENING_EXCLUSION_LABELS.get(
                    player_field(player, 'screening_exclusion_reason')
                    or participant.vars.get('screening_exclusion_reason'),
                    '—',
                ),
                profile_complete=complete,
                current_page=participant._current_page_name or '—',
                round_number=participant._round_number or '—',
                age_band=age_labels.get(age_key, '—'),
                gender=gender_labels.get(gender_key, '—'),
                last_request=participant._last_request_timestamp or 0,
            )
        )

    participant_rows.sort(key=lambda item: item['last_request'], reverse=True)
    eligible_players = [player for player in players if player_is_eligible(player)]
    screened_out_count = len(players) - len(eligible_players)
    profile_complete_count = sum(
        profile_is_complete(player) for player in eligible_players
    )

    return dict(
        generated_at=datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S'),
        is_official=is_official,
        collection_mode='RECOLECCIÓN OFICIAL' if is_official else 'DATOS DE PRUEBA',
        summary=dict(
            assigned=len(players),
            started=sum(bool(player.participant.visited) for player in players),
            screened_out=screened_out_count,
            profile_complete=profile_complete_count,
            reached_final_page=sum(
                player.participant._index_in_pages >= player.participant._max_page_index
                for player in players
            ),
        ),
        quota_dimensions=[
            build_quota_dimension(eligible_players, dimension, targets)
            for dimension in QUOTA_DIMENSIONS
        ],
        intersection=build_age_gender_intersection(eligible_players, targets),
        participant_rows=participant_rows,
        no_participants=not participant_rows,
        targets_configured=any(bool(value) for value in targets.values()),
    )
